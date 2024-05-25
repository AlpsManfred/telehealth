date_of_appointment = None

from openpyxl.styles import PatternFill
import telebot
from telebot import types
from cor_funcs import *
import openpyxl
from openpyxl.styles.colors import Color
import datetime

TOKEN = "6657152814:AAHvriYMy1XNAA6Glb5FT11kQbtYZfFY-Gk"
bot = telebot.TeleBot(TOKEN)


@bot.message_handler(commands=['start'])
def start_message(message):
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    button1 = types.KeyboardButton('ФОТО ВРАЧА')
    button3 = types.KeyboardButton('Записаться')
    button4 = types.KeyboardButton('Связаться')
    button5 = types.KeyboardButton('Посмотреть портфолио')
    button6 = types.KeyboardButton('Система лояльности')
    keyboard.add(button1, button3, button4, button5, button6)
    bot.send_message(message.chat.id, 'Приветствуем в нашем медицинском центре!', reply_markup=keyboard)


# Хэндлер на сообщения
@bot.message_handler(content_types=['text'])
def send_text(message):
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    back_button = types.KeyboardButton('Назад')
    keyboard.add(back_button)

    if message.text == 'ФОТО ВРАЧА':
        bot.send_photo(message.chat.id, open('doctor.jpg', 'rb'))
        bot.send_message(message.chat.id, 'ФИО: Доктор Хаус\nСтаж: 10 лет', reply_markup=keyboard)

    elif message.text == 'Записаться':
        callback_type = 'Дата'
        keyboard = types.InlineKeyboardMarkup()
        working_days = get_working_days_from_today()
        for working_day in working_days:
            working_day = working_day.strftime("%d.%m")
            #button = types.InlineKeyboardButton(str(working_day), callback_data=(working_day, callback_type, None))
            button = types.InlineKeyboardButton(str(working_day), callback_data=working_day)
            keyboard.add(button)
        bot.send_message(message.chat.id, 'Выберите дату для записи на прием:', reply_markup=keyboard)

    elif message.text == 'Связаться':
        tg, phone_number = get_tg_and_phone_number()
        bot.send_message(message.chat.id,
                         f'Контакт специалиста:\nТелефон: {phone_number}\ntg: {tg}',
                         reply_markup=keyboard)

    elif message.text == 'Посмотреть портфолио':
        bot.send_photo(message.chat.id, open('doctor.jpg', 'rb'))
        bot.send_message(message.chat.id,
                         "Доктор Иванов - высококвалифицированный специалист в области кардиологии с более чем 10-ти летним опытом работы."
                         " Он окончил с отличием Медицинский университет и прошел ординатуру в одной из ведущих клиник страны. Доктор Иванов "
                         "имеет сертификат специалиста по кардиологии и регулярно проходит обучение и стажировки для поддержания своей квалификации"
                         " на высоком уровне. Он работает в больнице Здоровье и принимает пациентов в своем кабинете. Доктор Иванов известен своим "
                         "профессионализмом, внимательным отношением к пациентам и успешными "
                         "результатами лечения. Его коллеги и пациенты высоко ценят его вклад "
                         "в развитие медицины и здоровья.\n", reply_markup=keyboard)

    elif message.text == 'Система лояльности':
        bot.send_message(message.chat.id, 'Система лояльности в разработке', reply_markup=keyboard)

    elif message.text == 'Назад':
        start_message(message)


@bot.callback_query_handler(func=lambda call: True)
def callback_inline(call):
    global date_of_appointment
    if '.' in call.data:
        date = call.data
        date_of_appointment = date
        keyboard = types.InlineKeyboardMarkup()
        workbook = openpyxl.load_workbook('telehealth.xlsx')
        worksheet = workbook["Лист1"]

        start_work_time = worksheet["D2"].value
        end_work_time = worksheet["E2"].value
        start_work_time = datetime.datetime.strptime(str(start_work_time), "%H:%M")
        end_work_time = datetime.datetime.strptime(str(end_work_time), "%H:%M")

        times = []
        while start_work_time <= end_work_time:
            times.append(start_work_time.strftime("%H:%M"))
            start_work_time += datetime.timedelta(minutes=30)

        worksheet = workbook[date]
        for cell in worksheet['2']:
            if cell.fill.fgColor.index == 4:
                continue
            else:
                a = worksheet[f"{cell.column_letter}1"].value
                if a:
                    a = a.strftime("%H:%M")
                    if a in times:
                        times.remove(str(a))

        for i in times:
            callback_type = 'Время'
            button1 = types.InlineKeyboardButton(i,
                                                 callback_data=f'{i}')
            keyboard.add(button1)
        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                              text='Выберите свободный слот для записи на прием:', reply_markup=keyboard)
    elif ':' in call.data:
        time = call.data
        keyboard = types.InlineKeyboardMarkup()
        workbook = openpyxl.load_workbook('telehealth.xlsx')
        worksheet = workbook[date_of_appointment]
        for cell in worksheet['1']:
            if cell.value:
                cell_time = cell.value.strftime("%H:%M")
                if cell_time == time:
                    worksheet[f'{cell.column_letter}2'].fill = PatternFill(patternType='solid', fgColor=Color(indexed=3))
                    workbook.save("telehealth.xlsx")
                    break
        bot.send_message(call.message.chat.id,
                         f'Вы успешно записаны на прием {date_of_appointment} в {time}.',
                         reply_markup=types.ReplyKeyboardMarkup(resize_keyboard=True,
                                                                        one_time_keyboard=True).add('Назад'))


bot.polling(none_stop=True)
